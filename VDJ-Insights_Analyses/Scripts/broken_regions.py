import json
import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.patches as mpatches
from openpyxl.styles import Alignment, Font


def open_json(data_path: str) -> dict:
    with open(data_path, "r") as f:
        return json.load(f)


def make_pivot_table(data: pd.DataFrame) -> pd.DataFrame:
    grouped = data.groupby(['Region', 'Sample'], as_index=False).agg({'N_contigs': 'sum'})
    return grouped.pivot(index='Region', columns='Sample', values='N_contigs')


def make_pivot_vdj(data: pd.DataFrame):
    pivot = pd.pivot_table(
        data,
        index=['Sample', 'Region'],
        columns='Segment',
        aggfunc='size',
        fill_value=0
    )
    return pivot


def evaluate_completeness(df: pd.DataFrame, species: str, margin: float):
    # define thresholds for each supported species
    if species == "homo_sapiens":
        thresholds = {
            #"IGH": {"V": 123, "D": 27, "J": 9},
            "IGH": {"V": 123, "D": 25, "J": 5},
            "IGK": {"V": 76,  "D": 0,  "J": 5},
            "IGL": {"V": 73,  "D": 0,  "J": 7},
            "TRA": {"V": 57,  "D": 3,  "J": 65},
            "TRB": {"V": 64,  "D": 2,  "J": 14},
            "TRG": {"V": 12,  "D": 0,  "J": 5},
        }
    elif species == "macaca_mulatta":
        thresholds = {
            "IGH": {"V": 184, "D": 46, "J": 7},
            "IGK": {"V": 116, "D": 0,  "J": 5},
            "IGL": {"V": 119, "D": 0,  "J": 8},
            "TRA": {"V": 57,  "D": 3,  "J": 65},
            "TRB": {"V": 77,  "D": 2,  "J": 14},
            "TRG": {"V": 12,  "D": 0,  "J": 5},
        }
    else:
        # make sure thresholds is always defined (or fail fast)
        raise ValueError(f"evaluate_completeness: unsupported species {species!r}")

    df = df.reset_index()

    def check_row(row):
        for segment, thr in thresholds[row["Region"]].items():
            if thr == 0:
                continue
            min_value = round(thr * (1 - margin))
            if row[segment] < min_value:
                return "Incomplete"
        return "Complete"

    df["Complete"] = df.apply(check_row, axis=1)
    return df



def ensure_all_regions(df: pd.DataFrame):
    all_samples = df['Sample'].unique()
    required_regions = {"IGH","IGK","IGL","TRA","TRB","TRG"}
    missing = []
    for s in all_samples:
        present = set(df[df['Sample']==s]['Region'])
        for r in required_regions - present:
            missing.append({'Sample': s, 'Region': r, 'V':0,'D':0,'J':0,'Complete':'Empty'})
    if missing:
        df = pd.concat([df, pd.DataFrame(missing)], ignore_index=True)
    return df

def export_region_function_counts(data_bcr: pd.DataFrame,
                                  data_tcr: pd.DataFrame,
                                  output_path: str, complete:pd.DataFrame):
    df = pd.concat([data_bcr, data_tcr], ignore_index=True)
    counts = (
        df
        .groupby(['Sample','Region','Function'])
        .size()
        .reset_index(name='Count'))

    pivot = (
        counts
        .pivot(index='Sample', columns=['Region','Function'], values='Count')
        .fillna(0).astype(int))

    func_order = ['Functional','ORF','Pseudo']
    regions = list(pivot.columns.levels[0])
    cols = pd.MultiIndex.from_product([regions, func_order], names=['Region','Function'])
    pivot = pivot.reindex(columns=cols, fill_value=0)

    if not output_path.lower().endswith('.xlsx'):
        raise ValueError("output_path must end with .xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pivot.to_excel(writer, sheet_name='Summary', startrow=0)
        complete.to_excel(writer, sheet_name='Completeness', startrow=0)
        ws = writer.sheets['Summary']

        ws.freeze_panes = 'B4'

        start_col = 2  # 1=A, 2=B, so Region headers start at column C (idx=2)
        for region in regions:
            span = len(func_order)
            first = start_col
            last  = start_col + span - 1
            top_left = ws.cell(row=1, column=first)
            bottom_right = ws.cell(row=1, column=last)
            ws.merge_cells(f"{top_left.coordinate}:{bottom_right.coordinate}")
            top_left.value = region
            top_left.alignment = Alignment(horizontal='center', vertical='center')
            top_left.font = Font(bold=True)
            start_col = last + 1
        for cell in ws['A']:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    return

def assembly_error_check(pivot_table: pd.DataFrame, pivot_vdj: pd.DataFrame):
    for _, row in pivot_vdj.iterrows():
        region = row["Region"]
        sample_id = row["Sample"]
        complete_status = row["Complete"]
        if complete_status == "Incomplete":
            if region in pivot_table.index and sample_id in pivot_table.columns:
                current_value = pivot_table.loc[region, sample_id]
                if current_value == 1:
                    pivot_table.loc[region, sample_id] = 3
                if current_value == 2:
                    pivot_table.loc[region, sample_id] = 4
        if complete_status == "Empty":
            if region in pivot_table.index and sample_id in pivot_table.columns:
                pivot_table.loc[region, sample_id] = 0
    return pivot_table

def make_comparison_plot(run_data, run_complete, regions, samples, output_dir, margin):
    """
    run_data and run_complete are dicts: { 'contigs': pivot_table_df, 'scaffolds': pivot_vdj_df }
    regions: list of regions in desired order
    samples: list of samples in desired order
    """
    fig, axes = plt.subplots(
        ncols=2,
        sharey=True,
        figsize=(len(regions)*0.2 + 5, len(samples)*0.2 + 4),
        constrained_layout=True
    )

    percentage = 100 - (margin * 100)

    # white colormap (we only draw text markers)
    cmap = LinearSegmentedColormap.from_list('white', [(1,1,1),(1,1,1)])

    for ax, run in zip(axes, ['contigs','scaffolds']):
        pt = run_data[run] \
            .reindex(index=regions, columns=samples, fill_value=0) \
            .copy()
        completeness = run_complete[run]
        pt = assembly_error_check(pt, completeness)
        arr = pt.values.T  # samples × regions

        ax.imshow(arr, cmap=cmap, aspect='auto')
        ax.set_xticks(np.arange(len(regions)))
        ax.set_xticklabels(regions, ha='center', fontsize=10)
        ax.xaxis.set_ticks_position('top')

        ax.set_yticks(np.arange(len(samples)))
        ax.set_yticklabels(samples, fontsize=10)

        twos_per_row = (pt == 2).sum(axis=1)
        four_per_row = (pt == 4).sum(axis=1)
        print(f"Fragmented regions in {run}:", twos_per_row + four_per_row)
        one_per_row = (pt == 1).sum(axis=1)
        print(f"Complete regions in {run}:\n", one_per_row)
        nul_per_row = (pt == 0).sum(axis=1)
        print(f"Missing regions in {run}:\n", nul_per_row)

        # text markers
        fontsize = 14
        for i in range(len(samples)):
            for j in range(len(regions)):
                value = arr[i, j]
                if value == 0: #No segments found
                    ax.text(j, i, "×", ha="center", va="center", color="red", fontweight='bold', fontsize=fontsize)
                elif value == 2: #Fragmented region but correct number of segments
                    ax.text(j, i, "-", ha="center", va="center", color="orange", fontweight='bold', fontsize=fontsize)
                elif value == 3: #Complete region but low number of segments
                    ax.text(j, i, "-", ha="center", va="center", color="orange", fontweight='bold', fontsize=fontsize)
                elif value == 4: #Fragmented region and low number of segments
                    ax.text(j, i, "-", ha="center", va="center", color="orange", fontweight='bold', fontsize=fontsize)
                else: #Complete region
                    ax.text(j, i, "✓", ha="center", va="center", color="green", fontweight='bold', fontsize=fontsize)
        ax.set_xlabel(f'Immune regions \n found in {run}', labelpad=10, fontsize=12)
        ax.xaxis.set_label_position('top')
        if ax is axes[0]:
            ax.set_ylabel('Samples', labelpad=10, fontsize=12)
    '''
    # shared legend
    legend_elements = [
        mpatches.Patch(color='green',  label=f'Successful >{percentage}% (✓ / -)'),
        mpatches.Patch(color='orange', label=f'Partial <{percentage}% (* / -)'),
        mpatches.Patch(color='red',    label='None (×)')
    ]
    axes[-1].legend(handles=legend_elements, loc='upper right',
                    bbox_to_anchor=(1.2, 1.0), borderaxespad=0.)
'''
    for idx, (ax, run) in enumerate(zip(axes, ['contigs', 'scaffolds'])):
        # for the right‐hand panel only:
        if idx == 1:
            # turn off the little tick‐marks and labels
            ax.tick_params(axis='y', which='both', length=0)

    pos0 = axes[0].get_position()  # Bbox for left heatmap
    pos1 = axes[1].get_position()
    center_x = pos0.x0 + (pos1.x1 - pos0.x0) / 2 + 0.08

    #fig.suptitle('Schematic representation of the completeness\nof extracted regions per haplotype', fontsize=16, x=center_x)
    #fig.subplots_adjust(top=0.85)

    out_svg = os.path.join(output_dir, "comparison_broken_regions.svg")
    out_pdf = os.path.join(output_dir, "comparison_broken_regions.pdf")
    plt.savefig(out_svg, bbox_inches='tight')
    plt.savefig(out_pdf, dpi=300, bbox_inches='tight')
    plt.close()


def main(path: str, species, margin):
    runs = ['contigs', 'scaffolds']
    fig_dir = os.path.join(path, 'figure')
    os.makedirs(fig_dir, exist_ok=True)
    output_dir = os.path.join(path, "figure", "broken_regions")
    os.makedirs(output_dir, exist_ok=True)

    pivot_table_run = {}
    pivot_complete_run = {}

    # determine sample & region ordering once
    sample_order = None
    region_order = ["IGH","IGK","IGL","TRA","TRB","TRG"]

    for run in runs:
        # --- read broken_regions JSONs ---
        j1 = open_json(os.path.join(path, run, "bcr", "broken_regions.json"))
        j2 = open_json(os.path.join(path, run, "tcr", "broken_regions.json"))
        # merge the two JSON dicts
        combined = {**j1}
        for sam, regs in j2.items():
            combined.setdefault(sam, {}).update({r: vs for r, vs in regs.items() if r not in combined[sam]})

        # --- build a small DataFrame for pivot_table ---
        rows = []
        for sample, genes in combined.items():
            for region, entries in genes.items():
                status = ("Fragmented" if any(entry.get("Extraction status") == "Fragmented" for entry in entries)
                          else "Complete" if any(entry.get("Extraction status") == "Complete" for entry in entries)
                else None)

                num_contigs = 2 if status == "Fragmented" else 1 if status == "Complete" or "Extracted" else None
                if species == "homo_sapiens":
                    rows.append({
                        'Sample': sample.replace(".fasta", ""),
                        'Region': region,
                        'N_contigs': num_contigs})
                else:
                    rows.append({
                        'Sample': sample.split("_",1)[0],
                        'Region': region,
                        'N_contigs': num_contigs})
        df_br = pd.DataFrame(rows)
        pivot_table_run[run] = make_pivot_table(df_br)

        # --- read annotation reports and build completeness table ---
        bcr = pd.read_excel(os.path.join(path, run, "bcr", "annotation", "annotation_report_all.xlsx"))
        tcr = pd.read_excel(os.path.join(path, run, "tcr", "annotation", "annotation_report_all.xlsx"))
        combined_vdj = pd.concat([make_pivot_vdj(bcr), make_pivot_vdj(tcr)], axis=0)
        #check for humans!
        if species == "macaca_mulatta":
            combined_vdj.rename(index=lambda x: x.split('_', 1)[0], inplace=True)

        complete = evaluate_completeness(combined_vdj, species, margin)
        complete = ensure_all_regions(complete)
        pivot_complete_run[run] = complete
        complete.to_excel(os.path.join(output_dir, f"{species}_{run}_completeness.xlsx"), index=False)

        # capture sample order from the first run
        if sample_order is None:
            sample_order = sorted(pivot_table_run[run].columns.tolist())

        output_path = os.path.join(output_dir, f"Functionality-count_{run}_{species}.xlsx")
        export_region_function_counts(bcr, tcr, output_path, complete)

    # now plot comparison
    make_comparison_plot(
        run_data=pivot_table_run,
        run_complete=pivot_complete_run,
        regions=region_order,
        samples=sample_order,
        output_dir=output_dir,
        margin=margin
    )


if __name__ == "__main__":
    species = "homo_sapiens"
    margin = 0.2
    main("/path/to/outcomes/HPGC/scaffolds/", species, margin)
